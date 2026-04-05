"""Excel 模板生成与批量导入（管理组、人员编号与组对应）。"""
from __future__ import annotations

import io
import re
from typing import Any

import openpyxl
from django.db import transaction
from openpyxl.styles import Font

from .models import PersonIdGroupAssignment, PhotoManagementGroup

# 模板与导入表头（首行）
GROUP_HEADERS = ('组名称', '代码', '说明')
ASSIGNMENT_HEADERS = ('证件号或人员编号', '管理组名称', '备注')


def _normalize_header(cell) -> str:
    if cell is None:
        return ''
    return re.sub(r'\s+', '', str(cell).strip())


def _match_headers(row_values: tuple, expected: tuple[str, ...]) -> bool:
    if len(row_values) < len(expected):
        return False
    got = tuple(_normalize_header(row_values[i]) for i in range(len(expected)))
    exp = tuple(re.sub(r'\s+', '', h) for h in expected)
    return got == exp


def build_group_template_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '管理组'
    bold = Font(bold=True)
    for col, title in enumerate(GROUP_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = bold
    ws.append(['示例学院', 'college-a', '可选说明文字'])
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40
    return wb


def build_assignment_template_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '人员与组'
    bold = Font(bold=True)
    for col, title in enumerate(ASSIGNMENT_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = bold
    ws.append(['20230001', '示例学院', '可选'])
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 30
    return wb


def workbook_to_response(wb: openpyxl.Workbook, filename: str) -> Any:
    from django.http import HttpResponse

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def import_groups_from_upload(upload_file) -> dict[str, Any]:
    """
    首行为表头：组名称、代码、说明。
    已存在的组名称会更新代码与说明。
    """
    wb = openpyxl.load_workbook(upload_file, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row or not _match_headers(header_row, GROUP_HEADERS):
        return {
            'ok': False,
            'error': f'表头须为：{"、".join(GROUP_HEADERS)}（勿改列名与顺序）',
        }

    created = updated = 0
    errors: list[str] = []

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            if not row or all(v is None or str(v).strip() == '' for v in row[:3]):
                continue
            name = row[0]
            code = row[1] if len(row) > 1 else ''
            desc = row[2] if len(row) > 2 else ''
            name_s = str(name).strip() if name is not None else ''
            if not name_s:
                errors.append(f'第 {idx} 行：组名称不能为空')
                continue
            code_s = str(code).strip() if code is not None else ''
            desc_s = str(desc).strip() if desc is not None else ''

            slug = code_s or ''
            if slug and not re.match(r'^[-a-zA-Z0-9_]+$', slug):
                errors.append(f'第 {idx} 行：代码仅允许英文、数字、连字符与下划线')
                continue

            obj, was_created = PhotoManagementGroup.objects.get_or_create(
                name=name_s,
                defaults={'code': slug, 'description': desc_s},
            )
            if was_created:
                created += 1
            else:
                obj.code = slug
                obj.description = desc_s
                obj.save(update_fields=['code', 'description'])
                updated += 1

    return {
        'ok': len(errors) == 0,
        'created': created,
        'updated': updated,
        'errors': errors,
    }


def import_assignments_from_upload(upload_file) -> dict[str, Any]:
    """
    首行：证件号或人员编号、管理组名称、备注。
    管理组名称须与「照片管理组」中名称完全一致。
    同一证件号重复导入将更新所属组与备注。
    """
    wb = openpyxl.load_workbook(upload_file, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row or not _match_headers(header_row, ASSIGNMENT_HEADERS):
        return {
            'ok': False,
            'error': f'表头须为：{"、".join(ASSIGNMENT_HEADERS)}（勿改列名与顺序）',
        }

    created = updated = 0
    errors: list[str] = []

    group_cache: dict[str, PhotoManagementGroup] = {}

    def get_group(name: str) -> PhotoManagementGroup | None:
        if name in group_cache:
            return group_cache[name]
        g = PhotoManagementGroup.objects.filter(name=name).first()
        if g:
            group_cache[name] = g
        return g

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            if not row or all(v is None or str(v).strip() == '' for v in row[:3]):
                continue
            pid = row[0]
            gname = row[1]
            remark = row[2] if len(row) > 2 else ''
            pid_s = str(pid).strip() if pid is not None else ''
            gname_s = str(gname).strip() if gname is not None else ''
            remark_s = str(remark).strip() if remark is not None else ''

            if not pid_s:
                errors.append(f'第 {idx} 行：证件号或人员编号不能为空')
                continue
            if not gname_s:
                errors.append(f'第 {idx} 行：管理组名称不能为空')
                continue

            mg = get_group(gname_s)
            if not mg:
                errors.append(f'第 {idx} 行：未找到管理组「{gname_s}」')
                continue

            obj, was_created = PersonIdGroupAssignment.objects.update_or_create(
                person_number=pid_s,
                defaults={
                    'management_group': mg,
                    'remark': remark_s,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

    return {
        'ok': len(errors) == 0,
        'created': created,
        'updated': updated,
        'errors': errors,
    }
